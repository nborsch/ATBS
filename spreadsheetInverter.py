#! python3
# -*- coding: utf-8 -*-
#
# Solution to Practice Project "Spreadsheet Cell Inverter"
# from "Automate The Boring Stuff", by Al Sweigart,
# Chapter 12.
#
# "Write a program to invert the row and column of the cells 
# in the spreadsheet. For example, the value at row 5, column 
# 3 will be at row 3, column 5 (and vice versa). This should be 
# done for all cells in the spreadsheet.
# 
# You can write this program by using nested for loops to read 
# in the spreadsheet’s data into a list of lists data structure. 
# This data structure could have sheetData[x][y] for the cell at 
# column x and row y. Then, when writing out the new spreadsheet, 
# use sheetData[y][x] for the cell at column x and row y."
# 
# Nadia Borsch      misc@nborsch.com        Jun/2018

import openpyxl, os

print(f"\n{'Spreadsheet Cell Inverter':>35}")
print(f"{'=========== ==== ========':>35}")

print(f"\nCurrent working directory is {os.getcwd()}.")

while True:
    workbook = input("Please enter the name of the spreadsheet file to invert:\n")
    if workbook.endswith(".xlsx"):
        break

# Set up workbook and worksheets
wb = openpyxl.load_workbook(workbook)
sheet = wb.active
new_sheet = wb.create_sheet(f"NEW {sheet.title}", 0)

# List of lists to store cell data
sheet_data = [[] for i in range(sheet.max_row)]

# Populating data structure with sheet data
for row in range(1, sheet.max_row + 1):
    for cell in range(1, sheet.max_column + 1):
        sheet_data[row - 1].append(sheet.cell(row=row, column=cell).value)

# Inverting data and saving to new sheet
for cell in range(1, sheet.max_column + 1):
    for row in range(1, sheet.max_row + 1):
        new_sheet.cell(row=cell, column=row).value = sheet_data[row - 1][cell - 1]

# Remove old sheet and rename new sheet
wb.remove(sheet)
new_sheet.title = sheet.title

# Save workbook
try:
    print("Saving...")
    wb.save(f"{workbook}")
    print("Done.")

except PermissionError:
    print("\nCOULD NOT SAVE FILE, PLEASE TRY AGAIN.")