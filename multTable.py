#! python3
# -*- coding: utf-8 -*-
#
# Solution to Practice Project "Multiplication Table Maker"
# from "Automate The Boring Stuff", by Al Sweigart,
# Chapter 12.
#
# "Create a program multiplicationTable.py that takes a 
# number N from the command line and creates an N × N 
# multiplication table in an Excel spreadsheet."
# 
# Nadia Borsch      misc@nborsch.com        Jun/2018

import openpyxl, os
from openpyxl.styles import Font

print(f"\n{'Multiplication Table Maker':>35}")
print(f"{'============== ===== =====':>35}")

# Get number from user
while True:
    try:
        number = int(input("\nEnter an integer for the multiplication table:\n"))
        if number:
            break
    except ValueError:
        continue

# Set up worksheet
print(f"\nCreating spreadsheet 'multTable.xlsx' in {os.getcwd()}...")
wb = openpyxl.Workbook()
sheet = wb["Sheet"]

# Populate worksheet with results
print("Populating multiplication table...")
for i in range(1, number + 1):

    # Insert row headers
    sheet.cell(row=i+1, column=1).value = i
    sheet.cell(row=i+1, column=1).font = Font(bold=True)
    
    for j in range(1, number + 1):
        if i == 1:
            # Insert column headers
            sheet.cell(row=i, column=j+1).value = j
            sheet.cell(row=i, column=j+1).font = Font(bold=True) 
        
        # Insert values
        sheet.cell(row=i+1, column=j+1).value = i*j

# Freeze headers
sheet.freeze_panes = 'A2'
sheet.freeze_panes = 'B2'

# Save spreadsheet
try:
    print("Saving spreadsheet...")
    wb.save("multTable.xlsx")

    print("Done.")
except PermissionError:
    print("\nCOULD NOT SAVE SPREADSHEET, PLEASE TRY AGAIN.")
