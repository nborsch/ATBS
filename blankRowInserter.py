#! python3
# -*- coding: utf-8 -*-
#
# Solution to Practice Project "Blank Row Inserter"
# from "Automate The Boring Stuff", by Al Sweigart,
# Chapter 12.
#
# "Create a program blankRowInserter.py that takes two 
# integers and a filename string as command line arguments. 
# Let’s call the first integer N and the second integer M. 
# Starting at row N, the program should insert M blank rows 
# into the spreadsheet."
# 
# Nadia Borsch      misc@nborsch.com        Jun/2018

import openpyxl, os

print(f"\n{'Blank Row Inserter':>35}")
print(f"{'===== === ========':>35}")

print(f"\nCurrent working directory is {os.getcwd()}.")

while True:
    workbook = input("Please enter the name of the spreadsheet file to which the rows will be inserted:\n")
    if workbook.endswith(".xlsx"):
        break

while True:
    try:
        rows = int(input("\nEnter the number of rows to be inserted:\n"))
        if rows:
            break
    except ValueError:
        continue

while True:
    try:
        start_from = int(input("\nEnter the row number from which rows will be inserted:\n"))
        if start_from:
            break
    except ValueError:
        continue

print(f"{rows} rows will be inserted in the active sheet in {workbook} starting from row {start_from}.\n")

# Set up workbooks and worksheets
wb          = openpyxl.load_workbook(workbook)
sheet       = wb.active
new_sheet   = wb.create_sheet(f"{sheet.title}", 0)

# Copy the content from rows above and including start_from
print("Working...")
for each_row in range(1, start_from + 1):
    for each_cell in range(1, sheet.max_column + 1):
        new_sheet.cell(row=each_row, column=each_cell).value = sheet.cell(row=each_row, column=each_cell).value

# Copy the content to the moved cells
for each_new_row in range(rows + start_from + 1, sheet.max_row + rows + start_from + 1):
    for each_new_cell in range(1, sheet.max_column + 1):
        new_sheet.cell(row=each_new_row, column=each_new_cell).value = sheet.cell(row=each_new_row - rows, column=each_new_cell).value

# Delete old worksheet
wb.remove(sheet)

# Save workbook
try:
    print("Saving...")
    wb.save(f"{workbook}")
    print("Done.")

except PermissionError:
    print("\nCOULD NOT SAVE SPREADSHEET, PLEASE TRY AGAIN.")
