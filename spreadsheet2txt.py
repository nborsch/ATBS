#! python3
# -*- coding: utf-8 -*-
#
# Solution to Practice Project "Spreadsheet to Text Files"
# from "Automate The Boring Stuff", by Al Sweigart,
# Chapter 12.
#
# "Write a program that performs the tasks of the previous 
# program in reverse order: The program should open a spreadsheet 
# and write the cells of column A into one text file, the cells 
# of column B into another text file, and so on."
# 
# Nadia Borsch      misc@nborsch.com        Jun/2018

import openpyxl, os
from openpyxl.utils import get_column_letter

print(f"\n{'Spreadsheet to Text Files':>40}")
print(f"{'~~~~~~~~~~~ ~~ ~~~~ ~~~~~':>40}")

# Get folder path from user
print(f"\nCurrent working directory is {os.getcwd()}:")
path = input("Please enter the desired folder path or leave blank to stay in current working directory.\n")
if path:
    os.chdir(path)

# Get file name from user
while True:
    spreadsheet = input("\nPlease enter the filename for the spreadsheet to be converted:\n")
    if spreadsheet.endswith(".xlsx"):
        break

print(f"\nThe data on the active sheet in {os.path.join(os.getcwd(), spreadsheet)} will be converted to text files.")

# Open workbook/worksheet
wb = openpyxl.load_workbook(spreadsheet)
sheet = wb.active

# Traverse spreadsheet
for column_count, column in enumerate(sheet.columns, 1):

    # Filename is first row in column
    filename = sheet.cell(row=1, column=column_count).value
    txt_file = open(f"{filename}.txt", "w")
    print(f"Saving column {get_column_letter(column_count)} to '{filename}.txt'...")
    
    for row in range(1, len(column) + 1):
        row_content = sheet.cell(row=row, column=column_count).value

        if row_content:
            txt_file.write(f"{row_content}\n")
        else:
            # Cell is empty, but there are still more rows to traverse
            txt_file.write("\n")
    
    txt_file.close()

print("Done.")